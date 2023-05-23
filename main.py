import openpyxl
from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor

bot = Bot(token='6023062210:AAFkUSbjhOy4V-4kRIIms6iYqM9KTgNivTA')
dp = Dispatcher(bot)


user_data = {} # Создаю словарь для хранения данных пользователей

 
@dp.message_handler(commands=['start'])# Добавляем команду /start
async def process_start_command(message: types.Message):
    await message.reply("Добро пожаловать! Как вас зовут?")
    
    user_data[message.chat.id] = {}# Сохраняем ID пользователя в словарь

@dp.message_handler(lambda message: message.chat.id in user_data and 'name' not in user_data[message.chat.id])
async def process_name(message: types.Message):
    user_data[message.chat.id]['name'] = message.text
    await bot.send_message(chat_id=message.chat.id, text="Выберите тип забега:",
                           reply_markup=types.ReplyKeyboardMarkup(
                               keyboard=[
                                   [types.KeyboardButton(text="Полумарафон")],
                                   [types.KeyboardButton(text="Марафон")],
                                   [types.KeyboardButton(text="Сверхмарафон")]
                               ],
                               resize_keyboard=True
                           ))


@dp.message_handler(lambda message: message.chat.id in user_data and 'race_type' not in user_data[message.chat.id]
                    and message.text in ["Полумарафон", "Марафон", "Сверхмарафон"])
async def process_race_type(message: types.Message):
    user_data[message.chat.id]['race_type'] = message.text
    #Работаем с файлом excel,добавляя в ячейки ФИО и тип забега
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "ФИО"
    ws['B1'] = "Тип Забега"
    row = [user_data[message.chat.id]['name'], user_data[message.chat.id]['race_type']]
    ws.append(row)
    wb.save(filename='registration.xlsx')

    
    reply_markup = types.ReplyKeyboardRemove()# Удаляю кнопки после регистрации
    await bot.send_message(chat_id=message.chat.id, text="Вы успешно зарегистрировались!",
                           reply_markup=reply_markup)

    
    del user_data[message.chat.id]# Удаляю данные пользователя из словаря

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)

